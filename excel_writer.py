"""
Módulo de escrita no Excel.

Responsável por inserir os dados extraídos na planilha Excel existente,
sempre na próxima linha vazia. Trata PermissionError de forma amigável.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config import COLUNAS, NOME_ABA, PLANILHA_EXCEL
from models import DadosComprovante


def inserir_no_excel(dados: DadosComprovante, caminho: Path = PLANILHA_EXCEL) -> bool:
    """
    Insere uma linha com os dados do comprovante na planilha Excel.

    Localiza a próxima linha vazia e escreve os 3 campos.
    Se a planilha não existir, cria uma nova com o cabeçalho correto.

    Args:
        dados: DadosComprovante com os 3 campos preenchidos.
        caminho: Caminho para a planilha Excel.

    Returns:
        True se a inserção foi bem-sucedida, False caso contrário.
    """
    try:
        # Se a planilha não existir, cria com cabeçalho
        if not caminho.exists():
            _criar_planilha_vazia(caminho)

        # Abre a planilha existente
        wb = load_workbook(str(caminho))

        # Tenta acessar a aba configurada; se não existir, usa a primeira
        if NOME_ABA in wb.sheetnames:
            ws = wb[NOME_ABA]
        else:
            ws = wb.active

        # Encontra a próxima linha vazia
        proxima_linha = ws.max_row + 1

        # Escreve os dados na linha (colunas A, B, C)
        ws.cell(row=proxima_linha, column=1, value=dados.nome)
        ws.cell(row=proxima_linha, column=2, value=dados.data)
        ws.cell(row=proxima_linha, column=3, value=dados.valor)

        wb.save(str(caminho))
        wb.close()

        return True

    except PermissionError:
        # ─────────────────────────────────────────────
        # Tratamento amigável: planilha aberta no Excel
        # ─────────────────────────────────────────────
        print("\n" + "=" * 60)
        print("⚠️  ERRO DE PERMISSÃO")
        print("=" * 60)
        print(f"Não foi possível salvar em: {caminho.name}")
        print()
        print("A planilha provavelmente está aberta no Excel.")
        print("Por favor:")
        print("  1. Feche o arquivo no Excel")
        print("  2. Execute o script novamente")
        print("=" * 60 + "\n")
        return False

    except Exception as e:
        print(f"\n❌ Erro inesperado ao salvar no Excel: {e}\n")
        return False


def _criar_planilha_vazia(caminho: Path) -> None:
    """Cria uma planilha vazia com o cabeçalho correto."""
    df = pd.DataFrame(columns=[COLUNAS["nome"], COLUNAS["data"], COLUNAS["valor"]])
    df.to_excel(str(caminho), index=False, sheet_name=NOME_ABA)
    print(f"📄 Planilha criada: {caminho.name}")
